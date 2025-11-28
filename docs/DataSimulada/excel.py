import pandas as pd
import numpy as np
import random
from datetime import datetime
import os
from openpyxl import load_workbook

# Configuración base
np.random.seed(42)
random.seed(42)

# Carpetas y plantilla
current_dir = os.getcwd()
output_dir = os.path.join(current_dir, "DtSimulations")
os.makedirs(output_dir, exist_ok=True)

plantilla_excel = os.path.join(current_dir, "DtSimulations", "Formato_InformeDeCostos.xlsx")

clientes_edp = {
    "Codelco Ventanas": 140000000,
    "Codelco Salvador": 0,
    "Anglo American": 30000000
}
empleados_mo = {
    "Codelco Ventanas": 120000000,
    "Codelco Salvador": 0,
    "Anglo American": 20000000
}

# Diccionario de categorías con descripción, rango de valores y unidad asociada
categorias = {
    'MO': {'nombre': 'Mano de obra', 'tipo': 'Egreso', 'unidad': 'hrs', 'rango': (50000, 150000)},
    'EPP': {'nombre': 'Elementos de protección personal', 'tipo': 'Egreso', 'unidad': 'par', 'rango': (30000, 120000)},
    'M': {'nombre': 'Material', 'tipo': 'Egreso', 'unidad': 'und', 'rango': (20000, 150000)},
    'H': {'nombre': 'Herramienta', 'tipo': 'Egreso', 'unidad': 'und', 'rango': (50000, 120000)},
    'GG': {'nombre': 'Gastos Generales', 'tipo': 'Egreso', 'unidad': 'mes', 'rango': (40000, 90000)}
}

# Listas de descripciones más realistas por categoría
descripciones = {
    'MO': ['Mano de obra instalación', 'Mano de obra limpieza', 'Mano de obra mantenimiento'],
    'EPP': [
        'Casco de seguridad',
        'Zapatos de seguridad',
        'Buzo ignífugo-antiácido',
        'Lentes de seguridad claro o in/out',
        'Protector auditivo 3M 1100',
        'Guantes de cuero cabritilla',
        'Respirador medio rostro 7500 3M',
        'Filtros 60923',
        'Arnés tipo paracaídas',
        'Colas para arnés (posicionamiento y con amortiguador de impacto)',
        'Cinta antitrauma',
        'Barbiquejo'
    ],
    'M': [
        'Lámparas halógenas recargables',
        'Brochas',
        'Escobilla de acero',
        'Paños',
        'Caja de herramientas',
        'Material de oficina',
        'Cables eléctricos'
    ],
    'H': ['Taladro eléctrico', 'Llave inglesa', 'Sierra circular'],
    'GG': ['Energía eléctrica', 'Agua potable', 'Internet empresarial']
}

# Parámetros para controlar balance
target_margin = 0.15   # margen objetivo sobre egresos (15% por defecto). Ajusta a 0.10 - 0.25 según quieras.
max_extra_edp = 3      # máximo EdP adicionales para balancear
# Probabilidades (weights) para elegir categoría; suma no necesita ser 1
category_weights = {
    'MO': 0.22,
    'EPP': 0.12,
    'M': 0.22,
    'H': 0.13,
    'GG': 0.13
}

# Crear 12 meses desde octubre 2024
start_date = datetime(2025, 11, 1)
months = pd.date_range(start=start_date, periods=1, freq='MS')

# Carpeta de salida
current_dir = os.getcwd()
output_dir = os.path.join(current_dir, "DtSimulations")
os.makedirs(output_dir, exist_ok=True)

# Función para generar fechas dentro del mes
def fechas_del_mes(mes):
    dias_en_mes = pd.date_range(start=mes, end=mes + pd.offsets.MonthEnd(0))
    return random.choices(dias_en_mes, k=random.randint(40, 80))

# Función para obtener descripción sin repetir
def obtener_descripcion(cat, usados):
    disponibles = [d for d in descripciones[cat] if d not in usados]
    if not disponibles:  # si se acabaron, reiniciar el ciclo
        usados.clear()
        disponibles = descripciones[cat][:]
    desc = random.choice(disponibles)
    usados.add(desc)
    return desc

# Función helper para elegir categoría con pesos
def elegir_categoria():
    keys = list(category_weights.keys())
    weights = list(category_weights.values())
    return random.choices(keys, weights=weights, k=1)[0]

# Generar y guardar archivos
for mes in months:
    registros = []
    num = 1
    
    # Diccionario para registrar qué descripciones ya se usaron por categoría
    usados_por_categoria = {k: set() for k in descripciones.keys()}
    
    fechas = fechas_del_mes(mes)
    ultimo_dia = mes + pd.offsets.MonthEnd(0)
    for fecha in sorted(fechas):
        # Seleccionar una categoría al azar
        cat = elegir_categoria()
        info = categorias[cat]
        
        # Obtener descripción sin repetir
        desc = obtener_descripcion(cat, usados_por_categoria[cat])

        cantidad = random.randint(1, 10)
        precio_unitario = round(random.uniform(*info['rango']), 0)
        total = round(precio_unitario * cantidad, 0)

        registros.append({
            'N°': num,
            'Fecha': fecha.strftime(format="%d/%m/%Y"),  # formato tipo 1/7/2025
            'Descripción': desc,
            'Categoria': cat,
            'Tipo': info['tipo'],
            'Cant.': cantidad,
            'Unidad': info['unidad'],
            'Precio unitario': int(precio_unitario),
        })
        num += 1
    for cliente, monto in clientes_edp.items():
        registros.append({
            'N°': num,
            'Fecha': f"{ultimo_dia.day}/{ultimo_dia.month:02d}/{ultimo_dia.year}",
            'Descripción': f"Pago Estado de Pago - {cliente}",
            'Categoria': 'EdP',
            'Tipo': 'Ingreso',
            'Cant.': 1,
            'Unidad': 'informe',
            'Precio unitario': monto
        })
        num += 1
    for empleado, monto in empleados_mo.items():
        registros.append({
            'N°': num,
            'Fecha': f"{ultimo_dia.day}/{ultimo_dia.month:02d}/{ultimo_dia.year}",
            'Descripción': f"Pago de Salarios - {empleado}",
            'Categoria': 'MO',
            'Tipo': 'Egreso',
            'Cant.': 1,
            'Unidad': 'informe',
            'Precio unitario': monto
        })
        num += 1

    # Cargar el archivo generado
    wb = load_workbook(plantilla_excel)
    ws = wb.active
    
    # Crear DataFrame del mes
    df_mes = pd.DataFrame(registros)

    # Posición donde empezar a escribir los datos
    start_row = 11
    start_col = 2  # columna B

    # Escribir encabezados
    for j, col_name in enumerate(df_mes.columns):
        ws.cell(row=start_row, column=start_col + j, value=col_name)

    # Escribir datos
    for i, row in df_mes.iterrows():
        for j, col_name in enumerate(df_mes.columns):
            ws.cell(row=start_row + 1 + i, column=start_col + j, value=row[col_name])

    # Guardar archivo
    file_name = os.path.join(output_dir, f"informe_costos_{mes.year}_{mes.month:02d}.xlsx")
    wb.save(file_name)

# Mostrar ejemplos
generated_files = [os.path.join(output_dir, f"informe_costos_{mes.year}_{mes.month:02d}.xlsx") for mes in months[:3]]
print("Archivos generados:", generated_files)
