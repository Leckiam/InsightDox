from openpyxl import load_workbook
import pandas as pd

def procesar_informe(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    # Determinar rango final (hasta la última fila no vacía)
    start_row = 12
    start_col = 2
    end_col = 10
    end_row = ws.max_row

    # Leer rango de interés
    data = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True):
        if all(v is None for v in row):
            continue
        data.append(row)
        
    columnas = ["N°", "Fecha", "Descripción", "Categoria", "Tipo", "Cantidad", "Unidad", "Precio unitario", "Total"]
    df = pd.DataFrame(data, columns=columnas)
    
    # Convertir columnas numéricas a enteros cuando corresponda
    for col in ['N°', 'Cantidad', 'Precio unitario', 'Total']:  # columnas numéricas
        df[col] = df[col].apply(lambda x: int(x) if pd.notnull(x) and x == int(x) else x)

    return df

def obtenerMesAnno(df):
    df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')

    mes = df['Fecha'].dt.month[0]
    anno = df['Fecha'].dt.year[0]
    return [mes,anno]

from .models import MovimientoEconomico, InformeCostos
from datetime import datetime

def obtener_naturaleza(categoria):
    if categoria == 'EdP':
        return 'VE'
    elif categoria == 'MO':
        return 'RE'
    else:
        return 'GA'

def cargar_movimientos_desde_df(df, informe):

    for _, row in df.iterrows():
        fecha = pd.to_datetime(row["Fecha"]).date()
        descripcion = str(row["Descripción"]).strip()
        categoria = row["Categoria"]
        naturaleza = obtener_naturaleza(categoria)
        cantidad = int(row["Cantidad"]) if not pd.isna(row["Cantidad"]) else 0
        unidad = str(row["Unidad"]).strip()
        precio_unitario = int(row["Precio unitario"]) if not pd.isna(row["Precio unitario"]) else 0

        mov, created = MovimientoEconomico.objects.get_or_create(
            fecha=fecha,
            descripcion=descripcion,
            defaults={
                "naturaleza": naturaleza,
                "categoria": categoria,
                "cantidad": cantidad,
                "unidad": unidad,
                "precio_unitario": precio_unitario,
                "informe": informe,
                "nro": int(row["N°"]) if "N°" in df.columns else 0,
            }
        )
        if not created:
            mov.cantidad = cantidad
            mov.precio_unitario = precio_unitario
            mov.unidad = unidad
            mov.categoria = categoria
            mov.naturaleza = naturaleza
            mov.save()
