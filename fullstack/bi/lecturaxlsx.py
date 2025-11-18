from openpyxl import load_workbook
import pandas as pd

def procesar_informe(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    # Determinar rango final
    start_row = 12
    start_col = 2
    end_col = 10
    end_row = ws.max_row

    # Leer rango de interés
    data = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True):
        if all(v is None for v in row):   # ignorar filas totalmente vacías
            continue
        data.append(row)

    columnas = ["N°", "Fecha", "Descripción", "Categoria", "Tipo",
                "Cantidad", "Unidad", "Precio unitario", "Total"]

    df = pd.DataFrame(data, columns=columnas)

    # ----------------------------------------
    # LIMPIEZA
    # ----------------------------------------

    # Convertir columnas numéricas a número (o NaN)
    cols_numericas = ["N°", "Cantidad", "Precio unitario", "Total"]
    df[cols_numericas] = df[cols_numericas].apply(pd.to_numeric, errors='coerce')

    # Convertir texto, eliminar espacios
    df["Descripción"] = df["Descripción"].astype(str).str.strip()
    df["Unidad"] = df["Unidad"].astype(str).str.strip()

    # 1) Eliminar filas que tienen solo 1 dato útil
    df = df[df.notna().sum(axis=1) > 1]

    # 2) Eliminar filas donde columnas obligatorias están vacías
    obligatorias = ["Fecha", "Descripción", "Categoria", "Cantidad", "Unidad", "Precio unitario", "N°"]
    df = df.dropna(subset=obligatorias)

    # 3) Eliminar filas donde los textos estén vacíos
    df = df[df["Descripción"].str.strip() != ""]
    df = df[df["Unidad"].str.strip() != ""]

    # 4) Convertir fecha correctamente
    df = df.dropna(subset=["Fecha"])  # eliminar fechas inválidas
    df["Fecha"] = df["Fecha"].apply(normalizar_fecha)

    # 5) Rellenar NaN numéricos con 0 y convertir a enteros
    for col in cols_numericas:
        df[col] = df[col].fillna(0).astype(int)

    # ----------------------------------------
    return df

def normalizar_fecha(valor):
    # Si ya es datetime → devolver directo
    if isinstance(valor, datetime):
        return valor

    # Si es string → forzar formato DD/MM/YYYY
    if isinstance(valor, str):
        valor = valor.strip().replace("-", "/")
        try:
            return datetime.strptime(valor, "%d/%m/%Y")
        except:
            try:
                return datetime.strptime(valor, "%Y/%m/%d")
            except:
                return None

    return None

def obtenerMesAnno(df):
    df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce')
    print(df['Fecha'][0])
    mes = int(df['Fecha'].dt.month[0])
    anno = int(df['Fecha'].dt.year[0])
    return [mes, anno]

from .models import MovimientoEconomico, InformeCostos
from datetime import datetime

def obtener_naturaleza(categoria):
    if categoria == 'EdP':
        return 'VE'
    elif categoria == 'MO':
        return 'RE'
    else:
        return 'GA'

def cargar_movimientos_desde_df(df, informe):
    print('Inicio carga movimientos')
    columnas_obligatorias = [
        "Fecha", "Descripción", "Categoria",
        "Cantidad", "Unidad", "Precio unitario", "N°"
    ]
    df = df.dropna(subset=columnas_obligatorias)
    df = df[df["Descripción"].str.strip() != ""]
    df = df[df["Unidad"].str.strip() != ""]
    df = df[df.notna().sum(axis=1) > 1]  # opcional
    for _, row in df.iterrows():
        try:
            # --- PROCESAR DATOS ---
            fecha = pd.to_datetime(row["Fecha"]).date()
            descripcion = str(row["Descripción"]).strip()
            categoria = row["Categoria"]
            naturaleza = obtener_naturaleza(categoria)
            cantidad = int(row["Cantidad"])
            unidad = str(row["Unidad"]).strip()
            precio_unitario = int(row["Precio unitario"])
            nro = int(row["N°"])

            print(f'Procesando: {descripcion} ({categoria})')
            mov, created = MovimientoEconomico.objects.get_or_create(
                fecha=fecha,
                descripcion=descripcion,
                defaults={
                    "naturaleza": naturaleza,
                    "categoria": categoria,
                    "cantidad": cantidad,
                    "unidad": unidad,
                    "precio_unitario": precio_unitario,
                    "informe": informe,
                    "nro": nro
                }
            )

            if not created:
                mov.cantidad = cantidad
                mov.precio_unitario = precio_unitario
                mov.unidad = unidad
                mov.categoria = categoria
                mov.naturaleza = naturaleza
                mov.save()

        except Exception as e:
            print(f"Error procesando fila N° {row['N°']}: {e}")